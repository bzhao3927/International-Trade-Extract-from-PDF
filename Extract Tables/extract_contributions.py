from agentic_doc.parse import parse
import json
import pandas as pd
from pathlib import Path
import re
from typing import List, Dict, Optional
import glob
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def clean_numeric_value(value: str) -> str:
    """
    Clean numeric value string by removing commas, spaces, and non-numeric characters except dot and minus.
    Returns cleaned string or empty string if no digits found.
    """
    if pd.isna(value):
        return ''
    # Convert to string and strip spaces
    s = str(value).strip()
    # Remove commas and spaces
    s = s.replace(',', '').replace(' ', '')
    # Keep only digits, dot, minus sign
    s = re.sub(r'[^\d\.-]', '', s)
    # If empty or only dot/minus, return empty
    if not re.search(r'\d', s):
        return ''
    return s

def save_parse_json(pdf_path: Path, json_output_dir: Path) -> Path:
    """Parse the PDF and save JSON output"""
    try:
        results = parse(str(pdf_path), result_save_dir=str(json_output_dir))
        json_path = Path(results[0].result_path)
        print(f"Saved parse JSON to {json_path}")
        return json_path
    except Exception as e:
        print(f"Error parsing {pdf_path}: {e}")
        return None

def extract_tables_from_json(json_file_path):
    """
    Extract tables from JSON file containing markdown with HTML tables
    
    Args:
        json_file_path (str): Path to the JSON file
    
    Returns:
        list: List of pandas DataFrames containing the extracted tables
    """
    with open(json_file_path, 'r', encoding='utf-8') as file:
        data = json.load(file)
    
    markdown_content = data.get('markdown', '')
    
    if not markdown_content:
        print("No 'markdown' key found in JSON file")
        return []
    
    table_pattern = r'<table[^>]*>.*?</table>'
    table_matches = re.findall(table_pattern, markdown_content, re.DOTALL | re.IGNORECASE)
    
    tables = []
    
    for i, table_html in enumerate(table_matches):
        try:
            soup = BeautifulSoup(table_html, 'html.parser')
            table = soup.find('table')
            
            if table:
                rows = []
                
                thead = table.find('thead')
                if thead:
                    header_rows = thead.find_all('tr')
                    for row in header_rows:
                        row_data = []
                        cells = row.find_all(['th', 'td'])
                        for cell in cells:
                            colspan = int(cell.get('colspan', 1))
                            cell_text = cell.get_text(strip=True)
                            row_data.append(cell_text)
                            for _ in range(colspan - 1):
                                row_data.append('')
                        rows.append(row_data)
                
                tbody = table.find('tbody')
                if tbody:
                    body_rows = tbody.find_all('tr')
                else:
                    body_rows = [tr for tr in table.find_all('tr') if tr.parent.name != 'thead']
                
                for row in body_rows:
                    row_data = []
                    cells = row.find_all(['th', 'td'])
                    for cell in cells:
                        colspan = int(cell.get('colspan', 1))
                        cell_text = cell.get_text(strip=True)
                        row_data.append(cell_text)
                        for _ in range(colspan - 1):
                            row_data.append('')
                    rows.append(row_data)
                
                if rows:
                    if len(rows) > 1:
                        df = pd.DataFrame(rows[1:], columns=rows[0])
                    else:
                        df = pd.DataFrame(rows)
                    
                    df = df.replace('', pd.NA)
                    
                    tables.append(df)
                    
                    print(f"Extracted Table {i+1}: {df.shape[0]} rows × {df.shape[1]} columns")
        
        except Exception as e:
            print(f"Error processing table {i+1}: {str(e)}")
            continue
    
    return tables

def extract_year_from_filename(filename: str) -> Optional[int]:
    match = re.search(r'(\d{4})', filename)
    return int(match.group(1)) if match else None

def find_contribution_columns(df: pd.DataFrame, year: int) -> Dict[str, str]:
    columns = df.columns.tolist()
    column_mapping = {}
    
    lower_columns = [col.lower() for col in columns]
    
    country_candidates = ['member state', 'country', 'member', 'state']
    for candidate in country_candidates:
        for i, col in enumerate(lower_columns):
            if candidate in col:
                column_mapping['country'] = columns[i]
                break
        if 'country' in column_mapping:
            break
    
    if year <= 2010:
        collection_candidates = ['collections', 'adjustments', 'collection', 'adjustment']
        for candidate in collection_candidates:
            for i, col in enumerate(lower_columns):
                if candidate in col and 'outstanding' not in col:
                    column_mapping['annual_contributions'] = columns[i]
                    break
            if 'annual_contributions' in column_mapping:
                break
        
        outstanding_candidates = ['outstanding', 'total outstanding']
        for candidate in outstanding_candidates:
            for i, col in enumerate(lower_columns):
                if candidate in col:
                    column_mapping['total_outstanding_contributions'] = columns[i]
                    break
            if 'total_outstanding_contributions' in column_mapping:
                break
    
    else:  # 2011-2016
        net_candidates = ['net contributions', 'net contribution', 'assessed contributions']
        for candidate in net_candidates:
            for i, col in enumerate(lower_columns):
                if candidate in col:
                    column_mapping['assessed_contributions'] = columns[i]
                    break
            if 'assessed_contributions' in column_mapping:
                break
    
    return column_mapping

def extract_contributions_data(tables: List[pd.DataFrame], year: int) -> List[Dict]:
    all_data = []
    
    for table_idx, df in enumerate(tables):
        if df.empty:
            continue
        
        column_mapping = find_contribution_columns(df, year)
        
        if 'country' not in column_mapping:
            print(f"No country column found in table {table_idx + 1}")
            print(f"Available columns: {list(df.columns)}")
            continue
        
        print(f"Processing table {table_idx + 1} with columns: {column_mapping}")
        
        for _, row in df.iterrows():
            country = row.get(column_mapping['country'], '')
            
            if pd.isna(country) or str(country).strip() == '':
                continue
            
            clean_country = str(country).strip()
            
            record = {
                'year': year,
                'country': clean_country,
                'annual_contributions': '',
                'total_outstanding_contributions': '',
                'assessed_contributions': ''
            }
            
            if year <= 2010:
                if 'annual_contributions' in column_mapping:
                    annual = row.get(column_mapping['annual_contributions'], '')
                    record['annual_contributions'] = clean_numeric_value(annual)
                
                if 'total_outstanding_contributions' in column_mapping:
                    outstanding = row.get(column_mapping['total_outstanding_contributions'], '')
                    record['total_outstanding_contributions'] = clean_numeric_value(outstanding)
                
                try:
                    annual_val = float(record['annual_contributions']) if record['annual_contributions'] else 0
                    outstanding_val = float(record['total_outstanding_contributions']) if record['total_outstanding_contributions'] else 0
                    if annual_val != 0 or outstanding_val != 0:
                        record['assessed_contributions'] = str(annual_val + outstanding_val)
                except:
                    pass
            
            else:  # 2011-2016
                if 'assessed_contributions' in column_mapping:
                    assessed = row.get(column_mapping['assessed_contributions'], '')
                    record['assessed_contributions'] = clean_numeric_value(assessed)
            
            all_data.append(record)
    
    return all_data

def save_tables_to_excel(tables: List[pd.DataFrame], output_file: str, year: int):
    if not tables:
        print("No tables to save")
        return
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for i, df in enumerate(tables):
            sheet_name = f'Table_{i+1}_{year}'
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            for row_num, row in enumerate(worksheet.iter_rows(min_row=1), 1):
                for cell in row:
                    cell.border = thin_border
                    if row_num > 1 and row_num % 2 == 0:
                        cell.fill = light_fill
    
    print(f"Tables saved to {output_file}")

def process_all_pdfs():
    docs_dir = Path("docs")
    json_dir = Path("json_outputs")
    excel_dir = Path("excel_outputs")
    
    json_dir.mkdir(exist_ok=True)
    excel_dir.mkdir(exist_ok=True)
    
    pdf_files = list(docs_dir.glob("*.pdf"))
    pdf_files.sort()
    
    print(f"Found {len(pdf_files)} PDF files to process")
    
    all_contributions_data = []
    
    for pdf_path in pdf_files:
        year = extract_year_from_filename(pdf_path.name)
        
        if not year or year < 2000 or year > 2016:
            print(f"Skipping {pdf_path.name} - year {year} not in range 2000-2016")
            continue
        
        print(f"\n--- Processing {pdf_path.name} (Year: {year}) ---")
        
        json_path = save_parse_json(pdf_path, json_dir)
        if not json_path:
            continue
        
        tables = extract_tables_from_json(json_path)
        if not tables:
            print(f"No tables found in {pdf_path.name}")
            continue
        
        excel_file = excel_dir / f"un_contributions_{year}.xlsx"
        save_tables_to_excel(tables, str(excel_file), year)
        
        contributions_data = extract_contributions_data(tables, year)
        all_contributions_data.extend(contributions_data)
        
        print(f"Extracted {len(contributions_data)} records from {pdf_path.name}")
    
    if all_contributions_data:
        merged_df = pd.DataFrame(all_contributions_data)
        merged_df = merged_df.sort_values(['year', 'country'])
        
        merged_file = Path("merged_contributions_2000_2016.xlsx")
        
        with pd.ExcelWriter(merged_file, engine="openpyxl") as writer:
            merged_df.to_excel(writer, index=False, sheet_name="All_Contributions")
            
            workbook = writer.book
            worksheet = writer.sheets["All_Contributions"]
            
            header_format = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for col_num, cell in enumerate(worksheet[1]):
                cell.fill = header_format
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"\n=== SUMMARY ===")
        print(f"Processed {len(pdf_files)} PDF files")
        print(f"Total records extracted: {len(all_contributions_data)}")
        print(f"Years covered: {sorted(merged_df['year'].unique())}")
        print(f"Countries: {len(merged_df['country'].unique())}")
        print(f"Merged file saved as: {merged_file}")
        
        print(f"\nSample of merged data:")
        print(merged_df.head(10).to_string(index=False))
    
    else:
        print("No contribution data found in any files")

def process_single_file_debug(pdf_filename: str = "2000.pdf"):
    docs_dir = Path("docs")
    json_dir = Path("json_outputs")
    excel_dir = Path("excel_outputs")
    
    json_dir.mkdir(exist_ok=True)
    excel_dir.mkdir(exist_ok=True)
    
    pdf_path = docs_dir / pdf_filename
    
    if not pdf_path.exists():
        print(f"PDF file not found: {pdf_path}")
        return
    
    year = extract_year_from_filename(pdf_path.name)
    print(f"Processing {pdf_path.name} (Year: {year})")
    
    json_path = save_parse_json(pdf_path, json_dir)
    if not json_path:
        return
    
    tables = extract_tables_from_json(json_path)
    
    if tables:
        print(f"Found {len(tables)} tables")
        
        excel_file = excel_dir / f"debug_{year}.xlsx"
        save_tables_to_excel(tables, str(excel_file), year)
        
        contributions_data = extract_contributions_data(tables, year)
        print(f"Extracted {len(contributions_data)} contribution records")
        
        if contributions_data:
            sample_df = pd.DataFrame(contributions_data)
            print("\nSample extracted data:")
            print(sample_df.head(10).to_string(index=False))
    else:
        print("No tables found")

if __name__ == "__main__":
    print("UN Contributions PDF Batch Processor")
    print("=" * 50)
    
    print("Step 1: Testing with single file (2000.pdf)...")
    process_single_file_debug("2000.pdf")
    
    print("\nDo you want to continue with full batch processing? (y/n): ")
    response = input().lower().strip()
    
    if response == 'y':
        print("\nStarting full batch processing...")
        process_all_pdfs()
    else:
        print("Debug complete. Check the output files in excel_outputs/ directory.")
